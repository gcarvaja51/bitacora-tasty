# -*- coding: utf-8 -*-
"""
Control de cambios por estrategia, en Excel.

Por que existe: los cambios estaban en git y en CLAUDE.md, pero no habia forma
de ATRIBUIR resultados a versiones. Cuando se pregunta "¿este ajuste mejoro o
empeoro?", sin esto la unica respuesta honesta es "no se puede saber". El
objetivo de llegar a 80% de win rate por familia necesita poder aprender de
cada iteracion, no solo iterar.

Genera un libro por estrategia en su carpeta correspondiente:
  estrategia direccional/  -> control_cambios_direccional.xlsx
  neutral/                 -> control_cambios_neutral.xlsx
  reversion a la media/    -> control_cambios_reversion.xlsx

Cada libro trae 3 hojas: Cambios, Escala de impacto, Seguimiento.

NIVELES DE IMPACTO (la escala no es decorativa: define que hay que revalidar)
  ALTO  - Cambia CUANDO se entra o CUANDO se sale. Gatillos, umbrales de score,
          TP/SL, ventanas horarias, frenos. Puede alterar el resultado de todos
          los trades siguientes. Exige medir win rate antes/despues.
  MEDIO - Cambia CUANTO o CON QUE. Sizing, seleccion de strikes, deltas,
          anchos, filtros secundarios. Afecta la magnitud del resultado, no la
          decision de entrar.
  BAJO  - No cambia ninguna decision de trading. Registro, reportes, P&L
          historico, visualizacion, correccion de datos.

Uso:
  python scripts/control_cambios.py            # regenera desde git
  python scripts/control_cambios.py --dry-run  # muestra sin escribir
"""
import subprocess, sys, os, re
from datetime import datetime

REPO = r"C:\Users\gcarv\bitacora-tasty"
MENTORIA = (r"C:\Users\gcarv\Documents\CARPETA PERSONAL\01. guillermo carvajal"
            r"\01_Sigma\mentoria alejandro")
BASE = os.path.join(MENTORIA, "estrategias automatizadas")

# Las carpetas se renumeraron (01_..05_) despues de la primera version de este
# script; apuntar a los nombres viejos no fallaba con error, creaba carpetas
# nuevas vacias al lado de las buenas y el libro real quedaba sin actualizar.
DESTINOS = {
    "DIRECCIONAL": (os.path.join(BASE, "02_estrategia direccional"), "control_cambios_direccional.xlsx"),
    "NEUTRAL":     (os.path.join(BASE, "03_neutral"),                "control_cambios_neutral.xlsx"),
    "REVERSION":   (os.path.join(BASE, "04_reversion a la media"),   "control_cambios_reversion.xlsx"),
    "RUEDA":       (os.path.join(BASE, "01_ciclo rueda"),            "control_cambios_rueda.xlsx"),
    # Las dos bitacoras tambien llevan control (pedido del usuario). No son
    # estrategias, pero un error ahi puede hacer que una estrategia PAREZCA
    # buena o mala sin serlo — se vio el 2026-08-03: 39 de 62 trades tenian el
    # P&L mal calculado por la reconciliacion, y eso invalido horas de analisis
    # sobre el gatillo de entrada. Por eso se registran, no se ignoran.
    "BITACORA_TASTY":   (os.path.join(MENTORIA, "manual bitacora tasty"), "control_cambios_bitacora_tasty.xlsx"),
    "BITACORA_TRADIER": (os.path.join(MENTORIA, "analisis tradier"),      "control_cambios_bitacora_tradier.xlsx"),
}

# Las bitacoras se clasifican por ARCHIVOS tocados, no por palabras: el mensaje
# de un commit dice QUE cambio funcionalmente, no en cual de los dos tableros.
# Los archivos si lo dicen sin ambiguedad.
ARCHIVOS_BITACORA = {
    "BITACORA_TASTY":   ["public/index.html", "src/metrics.js", "src/wheel.js",
                         "src/tastytrade.js", "src/bp_adapter"],
    "BITACORA_TRADIER": ["public/tradier.html", "src/metrics_tradier.js",
                         "src/positions_tradier_adapter.js", "src/tradier_closed_pnl_adapter.js",
                         "src/bp_tradier_adapter.js"],
}

# La clasificacion se hace en cascada y EXCLUYENTE, no por coincidencia suelta
# de palabras: un cambio de La Rueda no es un cambio de las estrategias de SPX,
# y un cambio de reportes no es un cambio de estrategia en absoluto. Sin esto
# el registro se llena de ruido y deja de servir para atribuir resultados.

# 1. Cambios que NO son de estrategia: interfaz, reportes, visualizacion.
#    Se excluyen por completo — no alteran ninguna decision de trading.
NO_ESTRATEGIA = ["bitacora", "historial", "reporte", "monitor spx", "grafico", "dashboard",
                 "columna", "etiqueta", "pestaña", "boton", "pwa", "service worker", "informe",
                 "watchlist", "screener de acciones", "notebooklm", "skill", "daemon", "sigma terminal",
                 "tradingview", "premercado", "curva de capital", "calendario"]

# 2. La Rueda es un pipeline aparte (Tradier, acciones/ETF, horizonte de semanas).
RUEDA_KW = ["rueda", "wheel", "csp", "covered call", "cash-secured", "asignacion", "roll"]

# 3. Estrategias de SPX.
CLAVES = {
    "DIRECCIONAL": ["direccional", "camino b", "camino a", "pullback", "tendencia",
                    "playbook", "minscore", "weinstein", "trailing", "stop tecnico", "fractal", "poc",
                    "score", "spx 0dte"],
    "NEUTRAL":     ["iron condor", "condor", "neutral", "gex", "gamma flip", "vanna", "muros", "dex"],
    "REVERSION":   ["reversion", "sma8", "sma 8", "alejamiento", "juez", "bisturi", "vela garcia",
                    "vela tiburon", "vela 9", "compas", "rsi"],
}
# 4. Ejecucion y contabilidad: afecta a las TRES de SPX y a la Rueda por igual.
TRANSVERSAL = ["p&l", "pnl", "reconciliacion", "comision", "fill", "deslizamiento", "worstnetprice",
               "decimal", "orden fantasma", "ejecucion", "slippage", "broker"]

ALTO = ["gatillo", "pullback", "minscore", "umbral", "score", "tp", "sl", "stop", "ventana",
        "entrada", "salida", "veto", "gate", "freno", "peso", "repesaje", "trailing", "norma", "roll",
        # cambios en los insumos de la decision: mover un check de temporalidad,
        # activar/desactivar un indicador o tocar la confluencia cambia que se
        # opera, aunque el mensaje no diga "umbral".
        "check", "confluencia", "fase", "2m a 5m", "15m a 5m", "de 2m", "de 15m",
        "reactivar", "desactivar", "indicador", "señal", "signal", "macd", "ema", "sma"]
MEDIO = ["sizing", "delta", "strike", "ancho", "contrato", "riesgo", "capital", "filtro",
         "comision", "fee", "cordura", "sanidad", "fair value", "margin"]


def clasificar_familias(msg, archivos=''):
    m = msg.lower()
    fams = set()
    for fam, rutas in ARCHIVOS_BITACORA.items():
        if any(r in archivos for r in rutas):
            fams.add(fam)

    if any(k in m for k in NO_ESTRATEGIA):
        return fams   # cuenta para las bitacoras, no para las estrategias
    if any(k in m for k in RUEDA_KW):
        return fams | {"RUEDA"}                       # la Rueda no contamina a SPX
    # Si el commit nombra explicitamente una estrategia, esa manda. Sin esto,
    # "Reversion a la Media: veto GEX..." caia tambien en NEUTRAL por la palabra
    # "gex", y el registro de cada estrategia se llenaba de cambios ajenos.
    EXPLICITO = [("REVERSION", ["reversion a la media", "alejamiento de sma"]),
                 ("DIRECCIONAL", ["spx direccional", "direccional:", "camino b", "camino a"]),
                 ("NEUTRAL", ["iron condor", "long put condor"])]
    for fam, marcas in EXPLICITO:
        if any(k in m for k in marcas):
            return fams | {fam}
    porTema = {f for f, kws in CLAVES.items() if any(k in m for k in kws)}
    if porTema:
        return fams | porTema
    if any(k in m for k in TRANSVERSAL):
        return fams | {"DIRECCIONAL", "NEUTRAL", "REVERSION", "RUEDA"}
    return fams


# Trailers opcionales en el mensaje del commit. Cuando estan, MANDAN sobre la
# heuristica — que solo mira el asunto y por eso puede errar en un commit que
# hace dos cosas. Caso real: 89941f9 ("La Rueda contaba dos veces...") arreglaba
# ademas el gate del roll, un cambio de decision de trading, y quedo como BAJO
# porque el asunto hablaba de contabilidad.
#
#   Impacto: ALTO|MEDIO|BAJO
#   Estrategia: DIRECCIONAL, RUEDA        (varias separadas por coma)
_TRAILER_IMPACTO = re.compile(r'^\s*impacto\s*:\s*(ALTO|MEDIO|BAJO)\s*$', re.I | re.M)
_TRAILER_FAMILIA = re.compile(r'^\s*estrategias?\s*:\s*(.+)$', re.I | re.M)
FAMILIAS_VALIDAS = {"DIRECCIONAL", "NEUTRAL", "REVERSION", "RUEDA",
                    "BITACORA_TASTY", "BITACORA_TRADIER"}


def impacto_declarado(cuerpo):
    m = _TRAILER_IMPACTO.search(cuerpo or "")
    return m.group(1).upper() if m else None


# Alias -> familia canonica. "TENDENCIA" es como se llama la familia en el codigo
# y en el strategy log; "DIRECCIONAL" es como se llama la carpeta. Los dos son lo
# mismo y los dos aparecen escritos en los commits.
_ALIAS_FAMILIA = {
    "TENDENCIA": "DIRECCIONAL",
    "DIRECCIONAL": "DIRECCIONAL",
    "REVERSION": "REVERSION",
    "NEUTRAL": "NEUTRAL",
    "RUEDA": "RUEDA",
    "BITACORA_TASTY": "BITACORA_TASTY",
    "BITACORA_TRADIER": "BITACORA_TRADIER",
}


def familias_declaradas(cuerpo):
    """Extrae las familias del trailer  TOLERANDO prosa alrededor.

    Antes se exigia el nombre limpio: se partia por comas y se comparaba entero
    contra FAMILIAS_VALIDAS. Cualquier explicacion en la misma linea rompia el
    match y el commit caia en silencio a la heuristica del asunto — con la misma
    pinta que un commit sin trailer, o sea sin ninguna señal de que algo fallo.

    Detectado el 2026-08-09: los 12 commits de esa sesion declararon la familia
    y NINGUNO se clasifico. "Estrategia: reversion a la media." daba
    REVERSION_A_LA_MEDIA. y "Estrategia: NEUTRAL 1DTE. El 0DTE no se toca" daba
    NEUTRAL_1DTE._EL_0DTE... Es el mismo tipo de falla que el resto de la sesion:
    una convencion que se cumple pero un parser que no la reconoce.

    Ahora se buscan los nombres conocidos como palabras dentro de la linea, asi
    que "TENDENCIA y NEUTRAL. REVERSION no se ve afectada" devuelve las tres.
    Eso ultimo es deliberado y hay que saberlo: si el commit NOMBRA una familia
    para decir que NO la toca, igual queda registrada. Es preferible un cambio de
    mas en el libro que uno de menos — el libro existe para no perder nada.
    """
    m = _TRAILER_FAMILIA.search(cuerpo or "")
    if not m:
        return None
    linea = m.group(1).upper()
    fams = set()
    for alias, canonica in _ALIAS_FAMILIA.items():
        patron = r"\b" + alias.replace("_", r"[_ ]") + r"\b"
        if re.search(patron, linea):
            fams.add(canonica)
    return fams or None


def clasificar_impacto(msg):
    m = msg.lower()
    # Un fix de datos/reporte no cambia decisiones aunque mencione palabras de alto impacto.
    if any(k in m for k in ["bitacora", "historial", "reporte", "monitor spx", "grafico",
                            "dashboard", "columna", "etiqueta", "pestaña", "boton"]):
        return "BAJO"
    if any(k in m for k in ALTO):
        return "ALTO"
    if any(k in m for k in MEDIO):
        return "MEDIO"
    return "BAJO"


def leer_git():
    out = subprocess.run(
        ["git", "log", "--format=%h|%ad|%an|%s", "--date=format:%Y-%m-%d|%H:%M", "--since=2026-07-01"],
        cwd=REPO, capture_output=True, text=True, encoding="utf-8", errors="replace")
    filas = []
    for linea in out.stdout.strip().split("\n"):
        p = linea.split("|")
        if len(p) < 5:
            continue
        sha, fecha, hora, autor, asunto = p[0], p[1], p[2], p[3], "|".join(p[4:])
        filas.append(dict(sha=sha, fecha=fecha, hora=hora, autor=autor, asunto=asunto))
    return filas


# Cache de las llamadas a git: sin esto cada sha se consultaba una vez por
# familia (6) y otra al escribir la fila, o sea ~7 subprocesos por commit. Con
# el hook de post-commit corriendo esto en cada commit, eso se nota.
_CACHE_CUERPO, _CACHE_ARCHIVOS = {}, {}


def cuerpo_commit(sha):
    if sha in _CACHE_CUERPO:
        return _CACHE_CUERPO[sha]
    out = subprocess.run(["git", "log", "-1", "--format=%b", sha], cwd=REPO,
                         capture_output=True, text=True, encoding="utf-8", errors="replace")
    txt = " ".join(l.strip() for l in out.stdout.strip().split("\n")
                   if l.strip() and not l.startswith("Co-Authored-By")
                   and not l.startswith("Claude-Session"))
    _CACHE_CUERPO[sha] = txt[:1200]
    return _CACHE_CUERPO[sha]


_CACHE_CRUDO = {}


def cuerpo_crudo(sha):
    """Cuerpo SIN aplanar. cuerpo_commit() une las lineas con espacios para que
    quepa en una celda, y eso rompe cualquier match anclado a linea: los
    trailers `Impacto:`/`Estrategia:` quedaban en medio del parrafo y no los
    veia nadie. Se detecto al estrenar la convencion — el propio commit que la
    introdujo no se clasifico."""
    if sha in _CACHE_CRUDO:
        return _CACHE_CRUDO[sha]
    out = subprocess.run(["git", "log", "-1", "--format=%b", sha], cwd=REPO,
                         capture_output=True, text=True, encoding="utf-8", errors="replace")
    _CACHE_CRUDO[sha] = out.stdout
    return _CACHE_CRUDO[sha]


def archivos_commit(sha):
    if sha in _CACHE_ARCHIVOS:
        return _CACHE_ARCHIVOS[sha]
    out = subprocess.run(["git", "show", "--stat", "--format=", "--name-only", sha], cwd=REPO,
                         capture_output=True, text=True, encoding="utf-8", errors="replace")
    _CACHE_ARCHIVOS[sha] = ", ".join([f for f in out.stdout.strip().split("\n") if f][:6])
    return _CACHE_ARCHIVOS[sha]


# ── Atribucion de trades a versiones ────────────────────────────────────────
#
# Las columnas "Trades bajo esta version" / "Win rate despues" / "Validado"
# estuvieron SIEMPRE vacias: el libro registraba el cambio pero nunca cerraba
# el ciclo, que es justo lo que hace falta para responder "¿esto mejoro o
# empeoro?". Se llenan siguiendo las reglas que la propia hoja Seguimiento ya
# declaraba:
#
#   1. Cada cambio de impacto ALTO abre un periodo. El periodo va desde ese
#      commit hasta el siguiente ALTO de la MISMA familia.
#   2. La huella de algoVersion separa por configuracion, pero NO se mueve con
#      un cambio de codigo (lo dice src/algo_version.js y se vio con
#      stopTimeframe el 2026-08-05). Por eso el periodo se delimita por fecha
#      de commit y la huella se reporta aparte, en la hoja Seguimiento.
#   3. Nada anterior al 2026-08-03 es comparable: 39 de 62 trades direccionales
#      tienen el P&L mal calculado por el /gainloss viejo, y son irrecuperables.
#   4. Muestra minima para concluir: 30 trades cerrados.
CORTE_PNL_FIABLE = "2026-08-03"
MUESTRA_MINIMA = 30
FAMILIA_A_EJECUCION = {"DIRECCIONAL": "TENDENCIA", "NEUTRAL": "NEUTRAL", "REVERSION": "REVERSION"}
PROD = "https://web-production-23473.up.railway.app"


def _get_json(url, timeout=25):
    import json as _json
    from urllib.request import urlopen
    with urlopen(url, timeout=timeout) as r:
        return _json.loads(r.read().decode("utf-8"))


def cargar_ejecuciones():
    """Trades reales. Produccion primero (es donde vive el volumen con los datos
    de verdad); si no hay red, cae a los archivos locales, que pueden estar
    viejos — y lo dice, en vez de reportar en silencio sobre datos rancios."""
    import json as _json
    try:
        spx = _get_json(f"{PROD}/api/tradier/executions").get("executions", [])
        w = _get_json(f"{PROD}/api/wheel-trading/executions")
        rueda = w if isinstance(w, list) else w.get("executions", [])
        return spx, rueda, "produccion"
    except Exception as e:
        print(f"   (sin acceso a produccion: {e} — se usan los archivos locales)")
        def _local(n):
            p = os.path.join(REPO, n)
            if not os.path.exists(p):
                return []
            with open(p, encoding="utf-8") as fh:
                d = _json.load(fh)
            return d if isinstance(d, list) else d.get("executions", [])
        return _local("tradier_executions.json"), _local("wheel_trading_executions.json"), "local (posiblemente desactualizado)"


def trades_de_familia(familia, spx, rueda):
    if familia == "RUEDA":
        return [e for e in rueda if e.get("phase") == "CERRADO"]
    fam = FAMILIA_A_EJECUCION.get(familia)
    if not fam:
        return []          # las dos bitacoras no son estrategias, no se atribuyen trades
    return [e for e in spx if e.get("strategyFamily") == fam]


def _abierto_en(e):
    return (e.get("timestamp") or e.get("filledAt") or "")[:16].replace("T", " ")


def resultado_oficial(e):
    """FASE 0 (2026-08-21): el numero oficial lo calcula el servidor.

    Este script no reimplementa la regla del dinero. `src/pnl_oficial.js` la
    aplica una sola vez y la manda en cada ejecucion como `resultadoOficial`, y
    aca solo se lee. Es lo que evita que vuelva el problema de origen: la misma
    pregunta con cinco respuestas distintas segun quien la hiciera.

    Devuelve (pnl, comparable, cuenta). `cuenta` es False para lo que ni siquiera
    fue una operacion (orden fantasma del sandbox): esas no entran a ningun lado.
    """
    r = e.get("resultadoOficial")
    if isinstance(r, dict):
        if r.get("fuente") == "no_operacion":
            return None, False, False
        if r.get("pendiente") or r.get("pnl") is None:
            return None, False, True
        return r["pnl"], bool(r.get("comparable")), True
    # Fallback: JSON local viejo o servidor sin desplegar. Se usa el numero del
    # broker y se marca NO comparable, que es lo que honestamente es.
    p = e.get("pnl")
    if isinstance(p, (int, float)):
        return p, False, True
    return None, False, True


def medir(trades, desde, hasta):
    """desde/hasta en 'YYYY-MM-DD HH:MM'. hasta=None -> abierto hasta hoy.

    Comparables y legado se cuentan SEPARADO. Mezclar lo medido contra la cadena
    real con lo medido contra los fills de Tradier daba promedios sin sentido:
    sobre los trades que tienen las dos mediciones, 4 de 12 cambian de signo.
    """
    dentro = [e for e in trades if _abierto_en(e) >= desde and (hasta is None or _abierto_en(e) < hasta)]
    comp, legado, descartados = [], [], 0
    for e in dentro:
        pnl, comparable, cuenta = resultado_oficial(e)
        if not cuenta:
            descartados += 1
            continue
        if pnl is None:
            continue
        (comp if comparable else legado).append(pnl)

    gan = [p for p in comp if p > 0]
    per = [p for p in comp if p <= 0]
    wr = round(len(gan) / len(comp) * 100) if comp else None
    return dict(
        n=len(dentro) - descartados,
        cerrados=len(comp),
        ganadores=len(gan),
        wr=wr,
        pnl=round(sum(comp), 2),
        # La variable que hay que vigilar en estructuras de credito: el win rate
        # engaña, lo que mata es el tamaño de la cola.
        perdida_media=round(sum(per) / len(per), 2) if per else None,
        ganancia_media=round(sum(gan) / len(gan), 2) if gan else None,
        legado=len(legado),
        legado_pnl=round(sum(legado), 2),
        descartados=descartados,
    )


def veredicto(m, desde):
    """`cerrados` son SOLO los medidos contra la cadena real. Un periodo con
    trades de legado y ninguno comparable no es 'sin trades': es un periodo que
    no se puede juzgar, y hay que decirlo con esas palabras."""
    legado = m.get("legado", 0)
    if m["cerrados"] == 0:
        if legado:
            return f"no comparable ({legado} trades medidos con Tradier)"
        return "sin trades aun"
    if desde[:10] < CORTE_PNL_FIABLE:
        return f"no comparable (P&L previo al {CORTE_PNL_FIABLE})"
    if m["cerrados"] < MUESTRA_MINIMA:
        return f"insuficiente ({m['cerrados']}/{MUESTRA_MINIMA})"
    return "SI"


def construir():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    commits = leer_git()
    print(f"commits leidos: {len(commits)}")
    dry = "--dry-run" in sys.argv
    spx, rueda, fuente = cargar_ejecuciones()
    print(f"trades para atribuir: {len(spx)} de SPX + {len(rueda)} de la Rueda  (fuente: {fuente})")

    COLORES = {"ALTO": "FFC7CE", "MEDIO": "FFEB9C", "BAJO": "C6EFCE"}

    for familia, (carpeta, archivo) in DESTINOS.items():
        filas = []
        for c in commits:
            crudo = cuerpo_crudo(c["sha"])          # trailers: hace falta el cuerpo sin aplanar
            fams = familias_declaradas(crudo) or clasificar_familias(c["asunto"], archivos_commit(c["sha"]))
            if familia not in fams:
                continue
            filas.append(dict(c, impacto=impacto_declarado(crudo) or clasificar_impacto(c["asunto"])))
        filas.sort(key=lambda r: (r["fecha"], r["hora"]), reverse=True)

        print(f"\n{familia}: {len(filas)} cambios  "
              f"(ALTO {sum(1 for f in filas if f['impacto']=='ALTO')}, "
              f"MEDIO {sum(1 for f in filas if f['impacto']=='MEDIO')}, "
              f"BAJO {sum(1 for f in filas if f['impacto']=='BAJO')})")
        if dry:
            for f in filas[:5]:
                print(f"   {f['fecha']} {f['hora']} [{f['impacto']:5}] {f['asunto'][:70]}")
            continue

        wb = Workbook()
        ws = wb.active
        ws.title = "Cambios"
        cols = ["Fecha", "Hora", "Impacto", "Qué cambió", "Por qué / detalle", "Quién lo pidió",
                "Dónde (archivos)", "Commit", "Trades bajo esta versión", "Win rate después", "Validado"]
        ws.append(cols)
        for i, col in enumerate(cols, 1):
            cel = ws.cell(1, i)
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="1F3864")
            cel.alignment = Alignment(vertical="center", wrap_text=True)
        # Periodos de medicion: cada ALTO abre uno y lo cierra el ALTO siguiente
        # de la misma familia (regla 1 de la hoja Seguimiento). `filas` viene en
        # orden descendente, asi que el "siguiente" cronologico es el ALTO que
        # aparece ANTES en la lista.
        trades = trades_de_familia(familia, spx, rueda)
        altos = [f for f in filas if f["impacto"] == "ALTO"]
        periodos = []
        for i, f in enumerate(altos):
            desde = f"{f['fecha']} {f['hora']}"
            hasta = f"{altos[i-1]['fecha']} {altos[i-1]['hora']}" if i > 0 else None
            m = medir(trades, desde, hasta)
            periodos.append(dict(sha=f["sha"], asunto=f["asunto"], desde=desde, hasta=hasta, **m))
        medicion = {p["sha"]: p for p in periodos}

        for f in filas:
            p = medicion.get(f["sha"])
            if p is None:
                # Solo los ALTO abren periodo; el resto no se mide, y decirlo es
                # mas util que dejar la celda en blanco (blanco se lee como
                # "falta llenar", y no es que falte: no aplica).
                col_i = col_j = col_k = "no abre periodo (solo ALTO)"
            elif not trades and familia.startswith("BITACORA"):
                col_i = col_j = col_k = "n/a — no es una estrategia"
            else:
                col_i = f"{p['n']} abiertos / {p['cerrados']} cerrados"
                col_j = f"{p['wr']}%  ({p['ganadores']}/{p['cerrados']})" if p["wr"] is not None else "sin datos"
                col_k = veredicto(p, p["desde"])
            ws.append([f["fecha"], f["hora"], f["impacto"], f["asunto"], cuerpo_commit(f["sha"]),
                       f["autor"], archivos_commit(f["sha"]), f["sha"], col_i, col_j, col_k])
            ws.cell(ws.max_row, 3).fill = PatternFill("solid", fgColor=COLORES[f["impacto"]])
            if p is not None:
                v = str(col_k)
                ws.cell(ws.max_row, 11).fill = PatternFill(
                    "solid", fgColor="C6EFCE" if v == "SI" else "FFEB9C" if v.startswith("insuf") else "F2F2F2")
        for col, ancho in zip("ABCDEFGHIJK", [11, 7, 9, 52, 80, 14, 34, 10, 22, 17, 11]):
            ws.column_dimensions[col].width = ancho
        for r in range(2, ws.max_row + 1):
            ws.cell(r, 4).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, 5).alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:K{ws.max_row}"

        ws2 = wb.create_sheet("Escala de impacto")
        for fila in [
            ["Nivel", "Qué significa", "Qué hay que hacer al aplicarlo"],
            ["ALTO", "Cambia CUÁNDO se entra o CUÁNDO se sale: gatillos, umbrales de score, "
                     "TP/SL, ventanas horarias, frenos.",
                     "Medir win rate antes y después. No cambiar dos cosas de impacto alto a la vez — "
                     "si se mueven juntas, no se puede saber cuál funcionó."],
            ["MEDIO", "Cambia CUÁNTO o CON QUÉ: sizing, selección de strikes, deltas, anchos, "
                      "filtros secundarios.",
                      "Revisar el P&L medio por trade. La decisión de entrar no cambia, "
                      "así que el win rate debería moverse poco."],
            ["BAJO", "No cambia ninguna decisión de trading: registro, reportes, P&L histórico, "
                     "visualización, corrección de datos.",
                     "No requiere revalidación de la estrategia. Pero si corrige P&L histórico, "
                     "invalida las comparaciones hechas antes de la corrección."],
        ]:
            ws2.append(fila)
        for i in range(1, 4):
            ws2.cell(1, i).font = Font(bold=True, color="FFFFFF")
            ws2.cell(1, i).fill = PatternFill("solid", fgColor="1F3864")
        for col, ancho in zip("ABC", [10, 62, 72]):
            ws2.column_dimensions[col].width = ancho
        for r in range(2, 5):
            for cc in range(1, 4):
                ws2.cell(r, cc).alignment = Alignment(wrap_text=True, vertical="top")
            ws2.cell(r, 1).fill = PatternFill("solid", fgColor=COLORES[ws2.cell(r, 1).value])

        ws3 = wb.create_sheet("Seguimiento")
        ws3.append(["Objetivo", "Win rate 80%"])
        ws3.append(["Ambito", familia])
        ws3.append([])
        ws3.append(["Cómo se llena", ""])
        ws3.append(["1", "Cada cambio de impacto ALTO abre un período nuevo de medición."])
        ws3.append(["2", "Los trades se atribuyen por huella de versión (campo algoVersion de cada "
                         "ejecución, ver src/algo_version.js) — no por fecha, porque un mismo día "
                         "puede tener dos versiones."])
        ws3.append(["3", "No se compara contra períodos anteriores al 2026-08-03: 39 de los 62 trades "
                         "direccionales previos tienen el P&L mal calculado por el /gainloss viejo, "
                         "y son irrecuperables."])
        ws3.append(["4", "Muestra mínima antes de concluir: 30 trades. Con menos, la diferencia "
                         "entre 60% y 80% no es distinguible del azar."])
        ws3.append([])
        # Trades / Ganadores / Win rate / P&L cuentan SOLO lo medido contra la
        # cadena real. "Legado" son los que solo tienen el numero de Tradier: se
        # muestran para no perder el historial, nunca se suman con los de al lado.
        ws3.append(["Período (commit)", "Qué cambió", "Desde", "Hasta", "Trades", "Ganadores",
                    "Win rate", "P&L", "Pérdida media", "Legado (Tradier)", "Validado"])
        for i in range(1, 12):
            ws3.cell(10, i).font = Font(bold=True, color="FFFFFF")
            ws3.cell(10, i).fill = PatternFill("solid", fgColor="1F3864")
        for p in periodos:
            ws3.append([p["sha"], p["asunto"], p["desde"], p["hasta"] or "(vigente)",
                        f"{p['n']} / {p['cerrados']} cerrados",
                        p["ganadores"],
                        f"{p['wr']}%" if p["wr"] is not None else "—",
                        p["pnl"],
                        p.get("perdida_media") if p.get("perdida_media") is not None else "—",
                        f"{p.get('legado', 0)}" if p.get("legado") else "—",
                        veredicto(p, p["desde"])])
            v = str(ws3.cell(ws3.max_row, 11).value)
            ws3.cell(ws3.max_row, 11).fill = PatternFill(
                "solid", fgColor="C6EFCE" if v == "SI" else "FFEB9C" if v.startswith("insuf") else "F2F2F2")
            ws3.cell(ws3.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")

        # Huellas de algoVersion observadas: separan por CONFIGURACION, que es
        # ortogonal a los periodos por commit (un cambio de codigo no mueve la
        # huella). Se listan aparte para poder cruzar las dos vistas.
        huellas = {}
        for e in trades_de_familia(familia, spx, rueda):
            hv = (e.get("algoVersion") or {}).get("huella")
            if not hv:
                continue
            pnl, comparable, cuenta = resultado_oficial(e)
            if not cuenta:
                continue          # orden fantasma: nunca hubo operacion
            d = huellas.setdefault(hv, dict(n=0, cerrados=0, gan=0, pnl=0.0,
                                            legado=0, desde="9999", hasta=""))
            d["n"] += 1
            t = _abierto_en(e)
            d["desde"], d["hasta"] = min(d["desde"], t), max(d["hasta"], t)
            if pnl is None:
                continue
            if not comparable:
                d["legado"] += 1  # medido con Tradier: se cuenta, no se suma
                continue
            d["cerrados"] += 1
            d["pnl"] += pnl
            if pnl > 0:
                d["gan"] += 1
        if huellas:
            ws3.append([])
            ws3.append(["Huella (algoVersion)", "= misma configuración", "Desde", "Hasta", "Trades",
                        "Ganadores", "Win rate", "P&L", "Pérdida media", "Legado (Tradier)", "Validado"])
            enc = ws3.max_row
            for i in range(1, 12):
                ws3.cell(enc, i).font = Font(bold=True, color="FFFFFF")
                ws3.cell(enc, i).fill = PatternFill("solid", fgColor="404040")
            for hv, d in sorted(huellas.items(), key=lambda kv: kv[1]["desde"]):
                wr = round(d["gan"] / d["cerrados"] * 100) if d["cerrados"] else None
                m = dict(cerrados=d["cerrados"], legado=d["legado"])
                ws3.append([hv, "", d["desde"], d["hasta"], f"{d['n']} / {d['cerrados']} cerrados",
                            d["gan"], f"{wr}%" if wr is not None else "—", round(d["pnl"], 2),
                            "—", f"{d['legado']}" if d["legado"] else "—",
                            veredicto(m, d["desde"])])

        ws3.append([])
        ws3.append(["Fuente de los trades", fuente])
        ws3.append(["Generado", datetime.now().strftime("%Y-%m-%d %H:%M")])
        for col, ancho in zip("ABCDEFGHIJK", [22, 60, 18, 18, 20, 11, 10, 11, 14, 16, 30]):
            ws3.column_dimensions[col].width = ancho
        for r in range(4, 9):
            ws3.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")

        os.makedirs(carpeta, exist_ok=True)
        destino = os.path.join(carpeta, archivo)
        try:
            wb.save(destino)
            print(f"   -> {destino}")
        except PermissionError:
            alt = destino.replace(".xlsx", "_nuevo.xlsx")
            wb.save(alt)
            print(f"   -> ABIERTO EN EXCEL, guardado como {alt}")


if __name__ == "__main__":
    construir()
